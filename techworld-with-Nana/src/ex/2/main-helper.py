"""
Write an app that reads a Spreadsheet file and do something with that data
- List each company with respective product count
- List products with inventory less than 10
- List each company with their total inventory value
- Write to Spreadsheet: Calculate and write inventory value for each product into spreadsheet; write this aditional value to a new column in the Spreadsheet doc
"""

"""
- io module -> create, read, write
- openpyxl -> python module made for working with Exel Spreadsheet
"""

import openpyxl

# in order to process this document we have to assign it to a value
inv_file = openpyxl.load_workbook("inventory.xlsx")
# the name of the page in the spreadsheet that we want to process
product_list = inv_file["Sheet1"]

products_per_supplier = {} # declare an empty dict
total_value_per_supplier = {}
products_under_10_inv = {}

# .max_row -> attribute
print(f"product list row number: {product_list.max_row}")

# for loop is for iterite through a list
for product_row in range(2, product_list.max_row + 1):
  supplier_name = product_list.cell(product_row, 4).value
  inventory = product_list.cell(product_row, 2).value
  price = product_list.cell(product_row, 3).value
  product_num = product_list.cell(product_row, 1).value
    # it could be done better by making it to write in the next empty cell and not n=5
  inventory_price = product_list.cell(product_row, 5) # we did not add .value cuz if we want to update the cell, we need the whole cell object
  
  # we have to make sure to not overwrite the count but to increase it when the same company occur
  ## calculation number of products per supplier
  if supplier_name in products_per_supplier:
    # append to dict and if present increase the counter - one-line way
    products_per_supplier[supplier_name] = products_per_supplier[supplier_name] + 1
    # append to dict and if present increase the counter - other way
      # current_num_products = products_per_supplier[supplier_name]
        # current_num_products = products_per_supplier.get(supplier_name # -> other way to access the value of the key from a dict
      # products_per_supplier[supplier_name] = current_num_products + 1
  else:
    products_per_supplier[supplier_name] = 1

  ## calculation total value of inventory per supplier
  if supplier_name in total_value_per_supplier:
    current_total_value = total_value_per_supplier.get(supplier_name)
    total_value_per_supplier[supplier_name] = current_total_value + inventory * price
  else:  
    total_value_per_supplier[supplier_name] = inventory * price

  ## logic products with inventory less than 10
  if inventory < 10:
    products_under_10_inv[int(product_num)] = int(inventory)

  ## add value for total inventory
  ### this data will be ephemeral, we need actualy to save. check openpyxl doc
  inventory_price.value = int(inventory * price)
# this save function does not overwrite the original file but it saves as new
inv_file.save("inventory_total_value.xlsx")

print(f"product per supplier: {products_per_supplier}")
print(f"total value per supplier: {total_value_per_supplier}")
print(f"products under 10 in inventory: {products_under_10_inv}")

